import csv
from io import BytesIO
from io import StringIO

from openpyxl import load_workbook

from catalogs.sheet_mapping_schema import (
    SHEET_MAPPING_CANONICAL_COLUMNS,
    SHEET_MAPPING_FIELD_ALIASES,
    SHEET_MAPPING_HEADER_DETECTION_GROUPS,
    SHEET_MAPPING_MODEL_FIELD_MAP,
)
from patterns.sheet_mapping_patterns import (
    CELL_WHITESPACE_PATTERN,
    HEADER_NON_ALPHANUMERIC_PATTERN,
    HEADER_WHITESPACE_PATTERN,
)
from src.idms_db2_phase2.domain.models import SheetMappingRow


class SheetMappingParser:
    """
    Parses Sheet Mapping files from CSV/text or XLSX.

    This class intentionally does not own:
    - Sheet Mapping column names
    - Sheet Mapping aliases
    - Header detection groups
    - Regex patterns

    Those belong to:
    - catalogs/sheet_mapping_schema.py
    - patterns/sheet_mapping_patterns.py
    """

    def __init__(self) -> None:
        self.diagnostics: list[str] = []

    def parse_uploaded_file(
        self,
        uploaded_file,
    ) -> list[SheetMappingRow]:
        self.diagnostics = []

        if uploaded_file is None:
            self.diagnostics.append("No Sheet Mapping file supplied.")
            return []

        file_name = str(uploaded_file.name or "").lower()
        raw_bytes = uploaded_file.getvalue()

        self.diagnostics.append(f"Sheet Mapping file name: {file_name}")
        self.diagnostics.append(f"Sheet Mapping file size bytes: {len(raw_bytes)}")

        if file_name.endswith(".xlsx"):
            return self.parse_xlsx_bytes(raw_bytes)

        if file_name.endswith(".xls"):
            self.diagnostics.append(
                "Unsupported .xls file detected. Save the file as .xlsx or .csv."
            )
            return []

        text = raw_bytes.decode(
            "utf-8-sig",
            errors="ignore",
        )

        self.diagnostics.append(f"CSV/text decoded length: {len(text)}")

        if text:
            sample = text[:500].replace("\r", "\\r").replace("\n", "\\n")
            self.diagnostics.append(f"CSV/text sample: {sample}")

        return self.parse_csv_text(text)

    def parse_csv_text(
        self,
        text: str,
    ) -> list[SheetMappingRow]:
        if not str(text or "").strip():
            self.diagnostics.append("CSV/text Sheet Mapping is empty.")
            return []

        stream = StringIO(text)
        reader = csv.reader(stream)
        raw_rows = [tuple(row) for row in reader]

        if not raw_rows:
            self.diagnostics.append("CSV/text Sheet Mapping has no rows.")
            return []

        headers = [
            self._cell_to_string(value)
            for value in raw_rows[0]
        ]

        self.diagnostics.append(f"CSV detected headers: {headers}")

        output: list[SheetMappingRow] = []

        for row in raw_rows[1:]:
            raw_row = self._row_to_dict(
                headers=headers,
                row=row,
            )

            mapping_row = self._to_mapping_row(raw_row)

            if self._has_useful_content(mapping_row):
                output.append(mapping_row)

        self.diagnostics.append(f"CSV parsed useful rows: {len(output)}")
        self._add_population_diagnostics(output)

        return output

    def parse_xlsx_bytes(
        self,
        raw_bytes: bytes,
    ) -> list[SheetMappingRow]:
        if not raw_bytes:
            self.diagnostics.append("XLSX Sheet Mapping is empty.")
            return []

        workbook = load_workbook(
            BytesIO(raw_bytes),
            data_only=True,
            read_only=True,
        )

        output: list[SheetMappingRow] = []

        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                continue

            header_index = self._find_header_row(
                rows=rows,
                sheet_title=worksheet.title,
            )

            if header_index < 0:
                self.diagnostics.append(
                    f"Sheet {worksheet.title}: no Sheet Mapping header row detected."
                )
                continue

            parsed_rows = self._parse_xlsx_rows_from_header(
                raw_rows=rows,
                header_index=header_index,
            )

            self.diagnostics.append(
                f"Sheet {worksheet.title}: parsed useful rows: {len(parsed_rows)}"
            )

            output.extend(parsed_rows)

        self.diagnostics.append(f"XLSX parsed useful rows: {len(output)}")
        self._add_population_diagnostics(output)

        return output

    def _parse_xlsx_rows_from_header(
        self,
        raw_rows: list[tuple],
        header_index: int,
    ) -> list[SheetMappingRow]:
        headers = [
            self._cell_to_string(value)
            for value in raw_rows[header_index]
        ]

        output: list[SheetMappingRow] = []

        for row in raw_rows[header_index + 1:]:
            raw_row = self._row_to_dict(
                headers=headers,
                row=row,
            )

            mapping_row = self._to_mapping_row(raw_row)

            if self._has_useful_content(mapping_row):
                output.append(mapping_row)

        return output

    def _row_to_dict(
        self,
        headers: list[str],
        row: tuple,
    ) -> dict[str, str]:
        raw_row: dict[str, str] = {}

        for index, header in enumerate(headers):
            clean_header = self._cell_to_string(header)

            if not clean_header:
                continue

            value = row[index] if index < len(row) else ""
            raw_row[clean_header] = self._cell_to_string(value)

        return raw_row

    def _find_header_row(
        self,
        rows: list[tuple],
        sheet_title: str,
    ) -> int:
        for index, row in enumerate(rows[:100]):
            normalized_cells = {
                self._normalize_header(self._cell_to_string(value))
                for value in row
                if value is not None
            }

            if index < 10:
                self.diagnostics.append(
                    f"Sheet {sheet_title}: row {index} normalized cells: "
                    f"{sorted(normalized_cells)}"
                )

            if self._normalize_header(SHEET_MAPPING_CANONICAL_COLUMNS[0]) in normalized_cells:
                return index

            if self._row_has_any_header(
                normalized_cells=normalized_cells,
                aliases=SHEET_MAPPING_HEADER_DETECTION_GROUPS[0],
            ):
                return index

            if (
                self._row_has_any_header(
                    normalized_cells=normalized_cells,
                    aliases=SHEET_MAPPING_HEADER_DETECTION_GROUPS[1],
                )
                and self._row_has_any_header(
                    normalized_cells=normalized_cells,
                    aliases=SHEET_MAPPING_HEADER_DETECTION_GROUPS[2],
                )
            ):
                return index

        return -1

    def _row_has_any_header(
        self,
        normalized_cells: set[str],
        aliases: list[str],
    ) -> bool:
        for alias in aliases:
            if self._normalize_header(alias) in normalized_cells:
                return True

        return False

    def _to_mapping_row(
        self,
        raw_row: dict[str, str],
    ) -> SheetMappingRow:
        values: dict[str, str] = {}

        for model_field, canonical_column in SHEET_MAPPING_MODEL_FIELD_MAP.items():
            values[model_field] = self._get(
                row=raw_row,
                canonical_name=canonical_column,
            )

        return SheetMappingRow(**values)

    def _get(
        self,
        row: dict[str, str],
        canonical_name: str,
    ) -> str:
        aliases = SHEET_MAPPING_FIELD_ALIASES.get(
            canonical_name,
            [canonical_name],
        )

        normalized_lookup = {
            self._normalize_header(key): value
            for key, value in row.items()
        }

        for alias in aliases:
            normalized_alias = self._normalize_header(alias)
            value = normalized_lookup.get(normalized_alias)

            if value is not None:
                return str(value).strip()

        return ""

    def _cell_to_string(
        self,
        value,
    ) -> str:
        if value is None:
            return ""

        text = str(value)
        text = text.replace("\ufeff", "")
        text = text.replace("\xa0", " ")
        text = text.replace("\r\n", "\n")
        text = text.replace("\r", "\n")
        text = CELL_WHITESPACE_PATTERN.sub(" ", text)

        return text.strip()

    def _normalize_header(
        self,
        value: str,
    ) -> str:
        text = self._cell_to_string(value)
        text = text.upper()
        text = text.replace("_", " ")
        text = HEADER_NON_ALPHANUMERIC_PATTERN.sub("", text)
        text = HEADER_WHITESPACE_PATTERN.sub(" ", text)

        return text.strip()

    def _has_useful_content(
        self,
        row: SheetMappingRow,
    ) -> bool:
        values = [
            getattr(row, field_name)
            for field_name in SHEET_MAPPING_MODEL_FIELD_MAP
        ]

        return any(str(value or "").strip() for value in values)

    def _add_population_diagnostics(
        self,
        rows: list[SheetMappingRow],
    ) -> None:
        record_count = sum(
            1
            for row in rows
            if str(row.cobol_record_idms or "").strip()
        )

        source_field_count = sum(
            1
            for row in rows
            if str(row.cobol_zone or "").strip()
            or str(row.reference_field_name_copybook or "").strip()
        )

        db2_table_count = sum(
            1
            for row in rows
            if str(row.new_db2_record or "").strip()
            or str(row.cross_application_db2_table or "").strip()
        )

        db2_column_count = sum(
            1
            for row in rows
            if str(row.new_db2_field_name or "").strip()
            or str(row.cross_application_db2_field_name or "").strip()
        )

        useful_context_count = sum(
            1
            for row in rows
            if (
                str(row.cobol_record_idms or "").strip()
                and (
                    str(row.cobol_zone or "").strip()
                    or str(row.reference_field_name_copybook or "").strip()
                )
                and (
                    str(row.new_db2_record or "").strip()
                    or str(row.cross_application_db2_table or "").strip()
                )
                and (
                    str(row.new_db2_field_name or "").strip()
                    or str(row.cross_application_db2_field_name or "").strip()
                )
            )
        )

        self.diagnostics.append(
            f"Sheet Mapping populated Cobol Record IDMS rows: {record_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping populated source field rows: {source_field_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping populated DB2 table rows: {db2_table_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping populated DB2 column rows: {db2_column_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping useful source-to-target rows: {useful_context_count}"
        )